import time
from selenium.webdriver import Chrome, ChromeOptions
from selenium.common.exceptions import NoSuchElementException
import openpyxl
import os
from os.path import abspath, dirname
import zipfile


def get_class_sheet(all_class_sheet_path, target_sheet_name, course):
    wb = openpyxl.load_workbook(all_class_sheet_path)
    all_sheet_names = wb.sheetnames
    print(all_sheet_names)
    for each_sheet_name in all_sheet_names:
        if target_sheet_name in each_sheet_name:
            target_sheet = wb[each_sheet_name]
            break
    else:
        print('目标表格全名为:' + each_sheet_name)
    if os.path.isfile(abspath(dirname(__file__))+r'\{}.xlsx'.format(target_sheet_name)):
        print("文件已经存在")
    else:
        class_list = []
        for row in target_sheet:
            # print(row[2].value)
            if row[2].value:
                class_list.append(row[2].value)
        else:
            new_wb = openpyxl.Workbook()
            sheet = new_wb.active
            sheet.title = 'Sheet1'
            for index in range(len(class_list)):
                if index:
                    sheet["A" + str(index + 1)] = class_list[index]
                    # sheet.append(['xxx',100])
                else:
                    sheet.append(['名字', '成绩', ' 签到缺勤', '签退缺勤', '期末考试客观题', '期末考试主观题', '讨论次数'])
            else:
                new_wb.save(abspath(dirname(__file__))+r'\{}.xlsx'.format(target_sheet_name + course))
                print("表格创建完毕")


def into_zjy():
    time.sleep(1)
    account_number = browser.find_element_by_css_selector('#login-tabs > div > div.ykt-tab-panel.ykt-active > div > div:nth-child(1) > div > input')
    account_number.clear()
    account_number.send_keys('10386')
    password = browser.find_element_by_css_selector('#x-modify > div > input.ykt-form-field.pwd-modify')
    password.clear()
    password.send_keys('200036@zjy')
    verification_code = input("input verification code:")
    browser.find_element_by_css_selector('#x-modify > div > input[type=text]:nth-child(4)').send_keys(verification_code)
    browser.find_element_by_css_selector('#btnLogin').click()
    time.sleep(2)


def switch_window():
    time.sleep(1)
    current_window = browser.current_window_handle
    browser.switch_to.window(current_window)
    time.sleep(1)


def find_class_schedule(course_name, class_name):
    switch_window()
    browser.find_element_by_css_selector('#cancel2 > span').click()
    # 班级变化需要调整标签路径
    courses = browser.find_elements_by_css_selector('#content_course > div > ul > li > div > div > span.kt-lesson-ut > a')
    for course in courses:
        if course_name in course.text:
            course.click()
            break
    switch_window()
    browser.find_elements_by_partial_link_text(class_name)[0].click()
    switch_window()
    browser.find_element_by_css_selector('#menu > ul > li:nth-child(4) > a').click()
    switch_window()
    browser.find_element_by_css_selector('#tabs > li:nth-child(2) > a').click()


def find_teach_day(until_what_month=3):
    print("----开始查询考勤等情况----")
    switch_window()
    school_days = browser.find_elements_by_css_selector('#content_container > div > div > ul.np-fts-cal-day > li > div.np-fts-li')
    for index in range(len(school_days)):
        school_days = browser.find_elements_by_css_selector('#content_container > div > div > ul.np-fts-cal-day > li > div.np-fts-li')
        try:
            school_days[index].click()
            print(school_days[index].text)
            browser.find_element_by_css_selector('#content_container > div > div > ul.np-fts-cal-day > li.np-fts-boxshadow.np-fts-cal-bg > div.np-fts-content > div.np-fts-ftcontent > div > h3 > a').click()
        except Exception:
            continue
        else:
            switch_window()
            find_msg()
    else:
        print(browser.find_element_by_xpath('//*[@id="faceTeachDate_month"]').text[-2:] + "录入完毕")
        if int(browser.find_element_by_xpath('//*[@id="faceTeachDate_month"]').text[-2]) > until_what_month:
            browser.find_element_by_css_selector('#content_container > div > div > div > a:nth-child(4)').click()
            find_teach_day(until_what_month)
        else:
            print("全部录入完毕")


def get_final_grade(select='set_grade', download_file_path=None):
    browser.find_element_by_css_selector('#menu > ul > li:nth-child(7) > a').click()
    switch_window()
    browser.find_element_by_css_selector('#maintab > li > div.np-hw-control > a.am-inline-block.am-margin-xs.am-text-warning').click()
    switch_window()
    if select == 'download_answer':
        download_exam_answer(download_file_path)
    else:
        names_list = browser.find_elements_by_xpath('//*[@id="container"]/div[1]/div/div[2]/ul/li/a/span[3]')
        grades_list = browser.find_elements_by_xpath('//*[@id="container"]/div[1]/div/div[2]/ul/li/a/span[1]')
        set_final_grade(names_list, grades_list)


def download_exam_answer(download_file_path):
    if os.path.exists(download_file_path):
        pass
    else:
        os.mkdir(download_file_path)
    name_buttons = browser.find_elements_by_css_selector('#container > div:nth-child(1) > div > div.np-left-list > ul > li > a > span.np-li-stuName')
    for name_button in name_buttons:
        name_button.click()
        time.sleep(1)
        try:
            browser.find_elements_by_xpath('//a[contains (@title,".zip") or contains(@title,".rar") or contains(@title,".7z")]')[-1].click()
            print(browser.find_elements_by_xpath('//a[contains (@title,".zip") or contains(@title,".rar") or contains(@title,".7z")]')[-1].get_property("title"), end='')
            time.sleep(1)
            if os.path.exists(download_file_path + r'\学生答案.rar'):
                os.rename(download_file_path + r'\学生答案.rar', download_file_path + r'\{}.zip'.format(name_button.text))
            elif os.path.exists(download_file_path + r'\学生答案.zip'):
                os.rename(download_file_path + r'\学生答案.zip', download_file_path + r'\{}.zip'.format(name_button.text))
            elif os.path.exists(download_file_path + r'\学生答案.7z'):
                os.rename(download_file_path + r'\学生答案.7z', download_file_path + r'\{}.zip'.format(name_button.text))
        except Exception:
            continue
        else:
            print("下载完毕")


def set_final_grade(names_list, grades_list):
    print('开始录入成绩')
    grade_column_index = 4
    wb = openpyxl.load_workbook(abspath(dirname(__file__)) + file_path)
    sheet = wb.active
    for row in sheet.rows:
        for index in range(len(names_list)):
            if row[0].value == names_list[index].text:
                row[grade_column_index].value = int(grades_list[index].text[:2])
    wb.save(abspath(dirname(__file__)) + file_path)
    print('完成录入成绩')


def find_msg():
    browser.find_element_by_css_selector('#table_content > div > div.am-panel-bd > div.np-activity-nav > a.np-nav-current').click()
    time.sleep(1)
    msgs = browser.find_elements_by_css_selector('#content_classCurrent > div > ul > li > div.np-hw-title > h2 > a')
    for index in range(len(msgs)):
        msgs = browser.find_elements_by_css_selector('#content_classCurrent > div > ul > li > div.np-hw-title > h2 > a')
        even_day_type = msgs[index].text[-2:]
        if even_day_type == '签到' or even_day_type == '签退':
            print(even_day_type + "缺勤:")
            msgs[index].click()
            switch_window()
            browser.find_element_by_css_selector('#scoremanager_tabshow > div > ul > li:nth-child(3) > a').click()
            switch_window()
            even_day_name_list = browser.find_elements_by_css_selector('td:nth-child(4)')
            get_msg(even_day_name_list, even_day_type)
            browser.back()
            time.sleep(1)
        elif even_day_type == '讨论':
            print("参与" + even_day_type + "名单：")
            msgs[index].click()
            switch_window()
            even_day_name_list = browser.find_elements_by_css_selector('#content_container > ul > li > div.np-question-avatar > a')
            time.sleep(1)
            get_msg(even_day_name_list, even_day_type)
    else:
        browser.back()
        browser.back()
        time.sleep(1)


def get_msg(even_day_name_list, even_day_type):
    name_list = []
    for even_name in even_day_name_list:
        name_list.append(even_name.text)
    print(name_list)
    set_msg(name_list, even_day_type)
    browser.back()
    time.sleep(1)


def set_msg(name_list, even_day_type):
    print('开始修改表格')
    if even_day_type == '签到':
        index = 2
    elif even_day_type == '签退':
        index = 3
    else:
        index = 6
    wb = openpyxl.load_workbook(abspath(dirname(__file__))+file_path)
    sheet = wb.active
    for row in sheet.rows:
        if row[0].value in name_list:
            if row[index].value is None:
                row[index].value = 0
            row[index].value = row[index].value + 1
    wb.save(abspath(dirname(__file__))+file_path)
    print('表格修改完毕')


if __name__ == '__main__':
    # get_class_sheet(abspath(dirname(__file__))+r'\18级-考勤表2020.1.6.xlsx', '18人工智能1', '微信小程序开发')
    zjy_url = "https://zjy2.icve.com.cn/portal/login.html"
    option = ChromeOptions()
    option.add_argument('--start-maximized')  # 必须全屏打开不然会找不到部分元素
    download_file_dir = r'\18ai1_exam_answer'
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': abspath(dirname(__file__)) + download_file_dir}
    option.add_experimental_option('prefs', prefs)
    browser = Chrome(options=option)
    browser.get(zjy_url)
    browser.implicitly_wait(3)  # 设置自动等待时间，每半秒查看一次，直到超过最大时长5
    into_zjy()   # 进入和登陆执教云
    find_class_schedule("微信小程序", "18人工智能1班")  # 开始查询对应课程与对应班级的课程表
    file_path = r'\18人工智能1-微信小程序开发.xlsx'  # 需要信息录入的文件名
    find_teach_day(3)   # 开始录入考勤等情况并写入表格
    # get_final_grade('set_grade')   # 开始录入客观题分数
    # get_final_grade('download_answer', abspath(dirname(__file__)) + download_file_dir)  # 开始下载主观题答案
    browser.quit()
